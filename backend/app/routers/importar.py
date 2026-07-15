"""
Carga masiva desde Excel.

  GET  /api/importar/plantilla → plantilla de cobranzas.
  POST /api/importar/cobranzas → crea deudor (por RUT) + cobranza por fila.
  GET  /api/importar/plantilla-gestiones → plantilla de gestiones.
  POST /api/importar/gestiones → registra gestiones en bloque sobre cobranzas
       existentes de un cliente (identificadas por su ID cliente).

Reglas:
  - El CLIENTE debe existir previamente (se busca por nombre de fantasía o
    razón social). La filial se busca dentro del cliente (opcional).
  - El deudor se busca por RUT: si ya existe se reutiliza (no se duplica).
  - Cada fila es independiente: las filas buenas entran aunque otras fallen.
  - Las gestiones cargadas en bloque quedan marcadas es_masivo=True y se
    atribuyen a la persona indicada en la fila (debe ser un usuario).
"""

from io import BytesIO
from uuid import UUID
from decimal import Decimal, InvalidOperation
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.database import get_db
from app.security import usuario_autorizado
from app.models.cliente import Cliente
from app.models.filial import Filial
from app.models.deudor import Deudor, ContactoDeudor
from app.models.cobranza import Cobranza
from app.models.gestion import Gestion
from app.models.usuario import Usuario


router = APIRouter(
    prefix="/api/importar",
    tags=["Carga masiva"],
    dependencies=[Depends(usuario_autorizado)],
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

COLUMNAS = [
    "RUT deudor*", "Nombre deudor*", "Teléfono", "Email",
    "Cliente*", "Filial", "ID cliente", "Monto deuda*",
    "Tipo documento (pagare/factura/letra/cheque/otro)", "N° documento",
    "Fecha atención (AAAA-MM-DD)", "Previsión", "Observaciones",
]

EJEMPLO = [
    "12345678-9", "Juan Pérez Soto", "+56 9 1234 5678", "juan@correo.cl",
    "Redsalud", "Valparaíso", "155001", 450000,
    "pagare", "PG-4521",
    "2026-05-12", "FONASA", "Ingresado por carga masiva",
]

TIPOS_DOCUMENTO = {"pagare", "factura", "letra", "cheque", "otro"}


@router.get("/plantilla")
def descargar_plantilla():
    """Plantilla Excel con las columnas esperadas y una fila de ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cobranzas"
    fill = PatternFill(start_color="3F3F46", end_color="3F3F46", fill_type="solid")
    for col, titulo in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = fill
    for col, valor in enumerate(EJEMPLO, start=1):
        ws.cell(row=2, column=col, value=valor)
    for col, ancho in enumerate([14, 26, 18, 24, 14, 14, 12, 12, 22, 14, 24, 12, 30], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="plantilla_carga_cobranzas.xlsx"'},
    )


class ResultadoImportacion(BaseModel):
    filas_procesadas: int
    cobranzas_creadas: int
    deudores_nuevos: int
    errores: list  # [{fila, error}]


def _texto(valor) -> str:
    return str(valor).strip() if valor is not None else ""


@router.post("/cobranzas", response_model=ResultadoImportacion)
async def importar_cobranzas(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Procesa el Excel de carga masiva fila por fila."""
    if not archivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser un Excel .xlsx (usa la plantilla).",
        )

    contenido = await archivo.read()
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se pudo leer el archivo. ¿Es un .xlsx válido?",
        )
    ws = wb.active

    creadas = 0
    deudores_nuevos = 0
    errores = []
    filas = 0

    # Cache de clientes para no consultar por cada fila.
    clientes = db.query(Cliente).all()

    def buscar_cliente(nombre: str):
        n = nombre.lower()
        for c in clientes:
            if (c.nombre_fantasia or "").lower() == n or c.razon_social.lower() == n:
                return c
        return None

    for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Saltar filas totalmente vacías.
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue
        filas += 1
        rut = _texto(fila[0])
        nombre = _texto(fila[1])
        telefono = _texto(fila[2])
        email = _texto(fila[3])
        nombre_cliente = _texto(fila[4])
        nombre_filial = _texto(fila[5])
        id_cliente = _texto(fila[6])
        monto_crudo = fila[7]
        tipo_documento = _texto(fila[8]).lower() if len(fila) > 8 else ""
        numero_documento = _texto(fila[9]) if len(fila) > 9 else ""
        fecha_atencion = fila[10] if len(fila) > 10 else None
        prevision = _texto(fila[11]) if len(fila) > 11 else ""
        observaciones = _texto(fila[12]) if len(fila) > 12 else ""

        try:
            # --- Validaciones ---
            if not rut or not nombre:
                raise ValueError("Falta RUT o nombre del deudor")
            if not nombre_cliente:
                raise ValueError("Falta el cliente")
            cliente = buscar_cliente(nombre_cliente)
            if cliente is None:
                raise ValueError(f"Cliente '{nombre_cliente}' no existe en el sistema")
            try:
                monto = Decimal(str(monto_crudo))
                if monto <= 0:
                    raise InvalidOperation
            except (InvalidOperation, TypeError):
                raise ValueError(f"Monto inválido: {monto_crudo!r}")
            if tipo_documento and tipo_documento not in TIPOS_DOCUMENTO:
                raise ValueError(
                    f"Tipo de documento inválido: '{tipo_documento}' "
                    f"(usar: {', '.join(sorted(TIPOS_DOCUMENTO))})"
                )

            filial = None
            if nombre_filial:
                filial = (
                    db.query(Filial)
                    .filter(Filial.cliente_id == cliente.id,
                            Filial.nombre.ilike(nombre_filial))
                    .first()
                )
                if filial is None:
                    raise ValueError(
                        f"Filial '{nombre_filial}' no existe para {nombre_cliente}"
                    )

            fecha = None
            if fecha_atencion:
                if isinstance(fecha_atencion, datetime):
                    fecha = fecha_atencion.date()
                elif isinstance(fecha_atencion, date):
                    fecha = fecha_atencion
                else:
                    fecha = date.fromisoformat(_texto(fecha_atencion))

            # --- Deudor: reutilizar por RUT o crear con sus contactos ---
            deudor = db.query(Deudor).filter(Deudor.rut == rut).first()
            if deudor is None:
                deudor = Deudor(rut=rut, nombre=nombre)
                contactos = []
                if telefono:
                    contactos.append(ContactoDeudor(tipo="celular", valor=telefono))
                if email:
                    contactos.append(ContactoDeudor(tipo="email", valor=email))
                deudor.contactos = contactos
                db.add(deudor)
                db.flush()
                deudores_nuevos += 1

            # --- Cobranza ---
            db.add(Cobranza(
                cliente_id=cliente.id,
                filial_id=filial.id if filial else None,
                deudor_id=deudor.id,
                id_clinica=id_cliente or None,
                monto_original=monto,
                monto_actual=monto,
                tipo_documento=tipo_documento or "pagare",
                numero_pagare=numero_documento or None,
                fecha_atencion=fecha,
                prevision=prevision or None,
                observaciones=observaciones or None,
            ))
            db.commit()
            creadas += 1
        except Exception as e:
            db.rollback()
            mensaje = str(e)
            if "uq_cobranza_clinica" in mensaje:
                mensaje = f"ID cliente '{id_cliente}' ya existe para {nombre_cliente}"
            errores.append({"fila": i, "error": mensaje[:200]})

    return ResultadoImportacion(
        filas_procesadas=filas,
        cobranzas_creadas=creadas,
        deudores_nuevos=deudores_nuevos,
        errores=errores,
    )


# ============================================================
# Carga masiva de GESTIONES
# ============================================================

COLUMNAS_GESTION = [
    "ID cliente*", "Fecha (AAAA-MM-DD)", "Gestión*", "Persona*",
]

EJEMPLO_GESTION = [
    "155001", "2026-07-01", "Llamada al deudor, se compromete a pagar el día 5.", "GRV",
]


@router.get("/plantilla-gestiones")
def descargar_plantilla_gestiones():
    """Plantilla Excel para cargar gestiones en bloque sobre un cliente."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gestiones"
    fill = PatternFill(start_color="3F3F46", end_color="3F3F46", fill_type="solid")
    for col, titulo in enumerate(COLUMNAS_GESTION, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = fill
    for col, valor in enumerate(EJEMPLO_GESTION, start=1):
        ws.cell(row=2, column=col, value=valor)
    for col, ancho in enumerate([16, 20, 60, 16], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="plantilla_carga_gestiones.xlsx"'},
    )


class ResultadoGestiones(BaseModel):
    filas_procesadas: int
    gestiones_creadas: int
    errores: list  # [{fila, error}]


@router.post("/gestiones", response_model=ResultadoGestiones)
async def importar_gestiones(
    cliente_id: UUID = Form(..., description="Cliente al que pertenecen los ID de la planilla"),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Registra gestiones en bloque sobre cobranzas existentes del cliente
    indicado. Cada fila trae el ID cliente de la cobranza, la fecha, el texto
    de la gestión y la persona que la realizó (debe ser un usuario). Las
    gestiones quedan marcadas como masivas (es_masivo=True).
    """
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if cliente is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="El cliente indicado no existe.",
        )
    if not archivo.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser un Excel .xlsx (usa la plantilla).",
        )

    contenido = await archivo.read()
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se pudo leer el archivo. ¿Es un .xlsx válido?",
        )
    ws = wb.active

    creadas = 0
    errores = []
    filas = 0
    # Cache de usuarios (por nombre en minúsculas) para no consultar por fila.
    usuarios = {u.nombre.lower(): u for u in db.query(Usuario).all()}

    for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue
        filas += 1
        id_cliente = _texto(fila[0])
        fecha_cruda = fila[1] if len(fila) > 1 else None
        gestion = _texto(fila[2]) if len(fila) > 2 else ""
        persona = _texto(fila[3]) if len(fila) > 3 else ""

        try:
            if not id_cliente:
                raise ValueError("Falta el ID cliente")
            if not gestion:
                raise ValueError("Falta el texto de la gestión")
            if not persona:
                raise ValueError("Falta la persona que realizó la gestión")

            cobranza = (
                db.query(Cobranza)
                .filter(Cobranza.cliente_id == cliente.id,
                        Cobranza.id_clinica == id_cliente)
                .first()
            )
            if cobranza is None:
                raise ValueError(
                    f"No existe una cobranza con ID cliente '{id_cliente}' en {cliente.razon_social}"
                )

            usuario = usuarios.get(persona.lower())
            if usuario is None:
                raise ValueError(f"La persona '{persona}' no es un usuario del sistema")

            fecha = None
            if fecha_cruda:
                if isinstance(fecha_cruda, datetime):
                    fecha = fecha_cruda
                elif isinstance(fecha_cruda, date):
                    fecha = datetime(fecha_cruda.year, fecha_cruda.month, fecha_cruda.day)
                else:
                    fecha = datetime.fromisoformat(_texto(fecha_cruda))

            gestion_obj = Gestion(
                cobranza_id=cobranza.id,
                usuario_id=usuario.id,
                descripcion=gestion,
                es_masivo=True,
            )
            if fecha is not None:
                gestion_obj.fecha_gestion = fecha
            db.add(gestion_obj)
            db.commit()
            creadas += 1
        except Exception as e:
            db.rollback()
            errores.append({"fila": i, "error": str(e)[:200]})

    return ResultadoGestiones(
        filas_procesadas=filas,
        gestiones_creadas=creadas,
        errores=errores,
    )
