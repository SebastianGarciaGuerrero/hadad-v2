"""
Endpoints de exportación a Excel (openpyxl).

  GET /api/exportar/gestiones → gestiones de los últimos N días (default 15),
      filtrable por cliente o por filial.
  GET /api/exportar/recupero  → detalle de pagos de un mes (reemplaza el
      Excel de recupero mensual). Usa la vista SQL vista_recupero.
  GET /api/exportar/rendicion → cuadro resumen por cliente/filial de un mes
      (lo que se envía a la clínica). Usa la vista SQL vista_rendicion.

Los archivos se generan en memoria (BytesIO): no se escribe nada a disco.
"""

import unicodedata
from io import BytesIO
from uuid import UUID
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.security import get_current_user
from app.models.gestion import Gestion, TipoGestion
from app.models.cobranza import Cobranza
from app.models.deudor import Deudor
from app.models.cliente import Cliente
from app.models.filial import Filial
from app.models.usuario import Usuario


# dependencies=[...] exige token válido en TODOS los endpoints del router.
router = APIRouter(
    prefix="/api/exportar",
    tags=["Exportar"],
    dependencies=[Depends(get_current_user)],
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ENCABEZADOS = [
    "Fecha gestión", "N° Hadad", "ID clínica", "RUT deudor", "Deudor",
    "Cliente", "Filial", "Tipo de gestión", "Descripción",
    "Próximo contacto", "Registrada por",
]

# Formato de montos: separador de miles, sin decimales (pesos chilenos).
FORMATO_PESOS = "#,##0"


def _hoja_con_encabezados(titulo: str, encabezados: list, anchos: list):
    """Crea un workbook con una hoja, encabezado con estilo y panel congelado."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita el nombre de hoja a 31 caracteres
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col, value=texto)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = fill
        celda.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    return wb, ws


def _responder_xlsx(wb: Workbook, nombre_archivo: str) -> StreamingResponse:
    """Serializa el workbook en memoria y lo devuelve como descarga."""
    # Los headers HTTP no aceptan bien acentos/ñ: normalizamos a ASCII
    # ("Valparaíso" → "Valparaiso") para que el nombre no se corrompa.
    nombre_archivo = (
        unicodedata.normalize("NFKD", nombre_archivo.replace(" ", "_"))
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


def _validar_cliente(db: Session, cliente_id: UUID) -> Cliente:
    """Devuelve el cliente o lanza 404."""
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Cliente con id {cliente_id} no encontrado",
        )
    return cliente


@router.get("/gestiones")
def exportar_gestiones(
    dias: int = Query(15, ge=1, le=365, description="Cuántos días hacia atrás incluir"),
    cliente_id: Optional[UUID] = Query(None, description="Filtrar por cliente (clínica)"),
    filial_id: Optional[int] = Query(None, description="Filtrar por filial (sucursal)"),
    db: Session = Depends(get_db),
):
    """
    Descarga un Excel con las gestiones de los últimos `dias` días.
    Filtros combinables: por cliente (toda la clínica) o por filial puntual.
    """
    desde = datetime.now(timezone.utc) - timedelta(days=dias)

    # Join: gestión → cobranza → deudor/cliente; filial y tipo pueden ser NULL
    # (outerjoin) para no perder gestiones de cobranzas sin filial asignada.
    query = (
        db.query(Gestion, Cobranza, Deudor, Cliente, Filial, TipoGestion, Usuario)
        .join(Cobranza, Gestion.cobranza_id == Cobranza.id)
        .join(Deudor, Cobranza.deudor_id == Deudor.id)
        .join(Cliente, Cobranza.cliente_id == Cliente.id)
        .outerjoin(Filial, Cobranza.filial_id == Filial.id)
        .outerjoin(TipoGestion, Gestion.tipo_id == TipoGestion.id)
        .join(Usuario, Gestion.usuario_id == Usuario.id)
        .filter(Gestion.fecha_gestion >= desde)
    )

    nombre_filtro = "todas"
    if cliente_id is not None:
        cliente = _validar_cliente(db, cliente_id)
        query = query.filter(Cobranza.cliente_id == cliente_id)
        nombre_filtro = (cliente.nombre_fantasia or cliente.razon_social)
    if filial_id is not None:
        filial = db.query(Filial).filter(Filial.id == filial_id).first()
        if not filial:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Filial con id {filial_id} no encontrada",
            )
        query = query.filter(Cobranza.filial_id == filial_id)
        nombre_filtro = f"{nombre_filtro}-{filial.nombre}" if cliente_id else filial.nombre

    filas = query.order_by(Gestion.fecha_gestion.desc()).all()

    # --- Armar el Excel ---
    anchos = [17, 10, 12, 13, 30, 16, 14, 22, 60, 16, 20]
    wb, ws = _hoja_con_encabezados(
        f"Gestiones últimos {dias} días", ENCABEZADOS, anchos
    )

    for i, (g, cob, deu, cli, fil, tipo, usu) in enumerate(filas, start=2):
        ws.cell(row=i, column=1, value=g.fecha_gestion.replace(tzinfo=None)
                ).number_format = "DD-MM-YYYY HH:MM"
        ws.cell(row=i, column=2, value=cob.numero)
        ws.cell(row=i, column=3, value=cob.id_clinica)
        ws.cell(row=i, column=4, value=deu.rut)
        ws.cell(row=i, column=5, value=deu.nombre)
        ws.cell(row=i, column=6, value=cli.nombre_fantasia or cli.razon_social)
        ws.cell(row=i, column=7, value=fil.nombre if fil else "")
        ws.cell(row=i, column=8, value=tipo.nombre if tipo else "")
        ws.cell(row=i, column=9, value=g.descripcion)
        if g.fecha_proximo_contacto:
            ws.cell(row=i, column=10, value=g.fecha_proximo_contacto
                    ).number_format = "DD-MM-YYYY"
        ws.cell(row=i, column=11, value=usu.nombre)

    hoy = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return _responder_xlsx(wb, f"gestiones_{nombre_filtro}_{dias}dias_{hoy}.xlsx")


@router.get("/recupero")
def exportar_recupero(
    anio: Optional[int] = Query(None, ge=2000, le=2100, description="Año (default: actual)"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Mes 1-12 (default: actual)"),
    cliente_id: Optional[UUID] = Query(None, description="Filtrar por cliente"),
    db: Session = Depends(get_db),
):
    """
    Descarga el RECUPERO del mes: el detalle de cada pago recibido, con su
    desglose capital/honorarios/interés. Reemplaza el Excel de recupero.
    Usa la vista SQL vista_recupero (siempre al día con los pagos).
    """
    ahora = datetime.now(timezone.utc)
    anio = anio or ahora.year
    mes = mes or ahora.month

    sql = """
        SELECT * FROM vista_recupero
        WHERE mes = make_date(:anio, :mes, 1)
    """
    params = {"anio": anio, "mes": mes}

    nombre_filtro = "todos"
    if cliente_id is not None:
        cliente = _validar_cliente(db, cliente_id)
        # La vista expone la razón social, no el id.
        sql += " AND cliente = :cliente"
        params["cliente"] = cliente.razon_social
        nombre_filtro = cliente.nombre_fantasia or cliente.razon_social

    sql += " ORDER BY fecha_pago, numero_cobranza"
    filas = db.execute(text(sql), params).mappings().all()

    encabezados = [
        "Fecha pago", "N° Hadad", "ID cliente", "Cliente", "Filial",
        "Deudor", "RUT deudor", "Total recibido", "Capital clínica",
        "Honorarios Hadad", "Interés clínica", "Gastos judiciales",
        "Estado", "Forma de pago", "N° comprobante", "Cuota", "Registrado por",
    ]
    anchos = [12, 10, 12, 16, 14, 30, 13, 14, 14, 15, 13, 15, 12, 14, 16, 8, 20]
    wb, ws = _hoja_con_encabezados(f"Recupero {mes:02d}-{anio}", encabezados, anchos)

    for i, r in enumerate(filas, start=2):
        ws.cell(row=i, column=1, value=r["fecha_pago"]).number_format = "DD-MM-YYYY"
        ws.cell(row=i, column=2, value=r["numero_cobranza"])
        ws.cell(row=i, column=3, value=r["id_clinica"])
        ws.cell(row=i, column=4, value=r["cliente"])
        ws.cell(row=i, column=5, value=r["filial"])
        ws.cell(row=i, column=6, value=r["deudor"])
        ws.cell(row=i, column=7, value=r["rut_deudor"])
        for col, campo in [(8, "total_recibido"), (9, "capital_clinica"),
                           (10, "honorarios_hadad"), (11, "interes_clinica"),
                           (12, "gastos_judiciales")]:
            ws.cell(row=i, column=col, value=r[campo]).number_format = FORMATO_PESOS
        ws.cell(row=i, column=13, value=r["estado_pago"])
        ws.cell(row=i, column=14, value=r["forma_pago"])
        ws.cell(row=i, column=15, value=r["numero_comprobante"])
        if r["numero_cuota"]:
            ws.cell(row=i, column=16,
                    value=f'{r["numero_cuota"]}/{r["total_cuotas_acuerdo"]}')
        ws.cell(row=i, column=17, value=r["registrado_por"])

    # Fila de totales al final (suma en Excel, se recalcula sola).
    if filas:
        tot = len(filas) + 2
        ws.cell(row=tot, column=7, value="TOTALES").font = Font(bold=True)
        for col in (8, 9, 10, 11, 12):
            letra = get_column_letter(col)
            celda = ws.cell(row=tot, column=col, value=f"=SUM({letra}2:{letra}{tot-1})")
            celda.font = Font(bold=True)
            celda.number_format = FORMATO_PESOS

    return _responder_xlsx(wb, f"recupero_{nombre_filtro}_{anio}-{mes:02d}.xlsx")


@router.get("/rendicion")
def exportar_rendicion(
    anio: Optional[int] = Query(None, ge=2000, le=2100, description="Año (default: actual)"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Mes 1-12 (default: actual)"),
    cliente_id: Optional[UUID] = Query(None, description="Filtrar por cliente"),
    db: Session = Depends(get_db),
):
    """
    Descarga el CUADRO DE RENDICIÓN del mes: resumen por cliente/filial de
    cuánto se rinde a la clínica (capital + interés) y cuánto queda en Hadad
    (honorarios). Es lo que se envía a la clínica cada mes.
    Usa la vista SQL vista_rendicion.
    """
    ahora = datetime.now(timezone.utc)
    anio = anio or ahora.year
    mes = mes or ahora.month

    sql = """
        SELECT * FROM vista_rendicion
        WHERE mes = make_date(:anio, :mes, 1)
    """
    params = {"anio": anio, "mes": mes}

    nombre_filtro = "todos"
    if cliente_id is not None:
        cliente = _validar_cliente(db, cliente_id)
        sql += " AND cliente = :cliente"
        params["cliente"] = cliente.razon_social
        nombre_filtro = cliente.nombre_fantasia or cliente.razon_social

    sql += " ORDER BY cliente, filial"
    filas = db.execute(text(sql), params).mappings().all()

    encabezados = [
        "Cliente", "Filial", "Cantidad de pagos", "Capital clínica",
        "Interés clínica", "Total a rendir a clínica",
        "Honorarios Hadad", "Total recibido",
    ]
    anchos = [20, 16, 16, 15, 14, 20, 16, 14]
    wb, ws = _hoja_con_encabezados(f"Rendición {mes:02d}-{anio}", encabezados, anchos)

    for i, r in enumerate(filas, start=2):
        ws.cell(row=i, column=1, value=r["cliente"])
        ws.cell(row=i, column=2, value=r["filial"])
        ws.cell(row=i, column=3, value=r["cantidad_pagos"])
        for col, campo in [(4, "total_capital_clinica"), (5, "total_interes_clinica"),
                           (6, "total_a_rendir_clinica"), (7, "total_honorarios_hadad"),
                           (8, "total_recibido")]:
            ws.cell(row=i, column=col, value=r[campo]).number_format = FORMATO_PESOS

    if filas:
        tot = len(filas) + 2
        ws.cell(row=tot, column=2, value="TOTALES").font = Font(bold=True)
        celda = ws.cell(row=tot, column=3, value=f"=SUM(C2:C{tot-1})")
        celda.font = Font(bold=True)
        for col in (4, 5, 6, 7, 8):
            letra = get_column_letter(col)
            celda = ws.cell(row=tot, column=col, value=f"=SUM({letra}2:{letra}{tot-1})")
            celda.font = Font(bold=True)
            celda.number_format = FORMATO_PESOS

    return _responder_xlsx(wb, f"rendicion_{nombre_filtro}_{anio}-{mes:02d}.xlsx")
